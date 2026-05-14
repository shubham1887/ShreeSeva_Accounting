"""
ShreeSeva Medical Accounting Reconciliation Script
Takeover Date: 29-May-2025

Steps:
1. Read and join 3 DBF tables (ACCOUNT, PAIDMAST, PAIDTRAN)
2. Update Old_Bank_Statement: add Voucher No. and Party Name columns
3. Update System_Purchase: fill CHQ No., Date, Type, Previous Pending
4. Export joined DBF data to Excel
"""

import struct
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date
from collections import defaultdict

TAKEOVER_DATE = "20250529"
TAKEOVER_DT = date(2025, 5, 29)


# ─── DBF Reader ──────────────────────────────────────────────────────────────

def read_dbf(filename):
    with open(filename, "rb") as f:
        header = f.read(32)
        num_records = struct.unpack("<I", header[4:8])[0]
        header_size = struct.unpack("<H", header[8:10])[0]
        record_size = struct.unpack("<H", header[10:12])[0]

        fields = []
        while True:
            field_data = f.read(32)
            if field_data[0] == 0x0D:
                break
            name = field_data[:11].replace(b"\x00", b"").decode("latin-1")
            ftype = chr(field_data[11])
            flen = field_data[16]
            fields.append((name, ftype, flen))

        f.seek(header_size)
        records = []
        for _ in range(num_records):
            rec = f.read(record_size)
            if not rec or rec[0] == 0x1A:
                break
            if rec[0] == ord("*"):
                continue  # deleted record
            row = {}
            pos = 1
            for name, ftype, flen in fields:
                val = rec[pos : pos + flen].decode("latin-1").strip()
                row[name] = val
                pos += flen
            records.append(row)
    return records


def parse_date(s):
    """Convert YYYYMMDD string to date object."""
    if s and len(s) == 8 and s.isdigit():
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            pass
    return None


def fmt_date(d):
    """Format date as DD-Mon-YY string."""
    if d:
        return d.strftime("%d-%b-%y")
    return ""


# ─── Step 1: Load & Join DBF Tables ──────────────────────────────────────────

print("=" * 60)
print("Step 1: Loading DBF tables...")

account  = read_dbf("DataBase/ACCOUNT.DBF")
paidmast = read_dbf("DataBase/PAIDMAST.DBF")
paidtran = read_dbf("DataBase/PAIDTRAN.DBF")

print(f"  ACCOUNT  : {len(account)} records")
print(f"  PAIDMAST : {len(paidmast)} records")
print(f"  PAIDTRAN : {len(paidtran)} records")

# Build lookup maps
acct_by_id   = {r["ACCOID"]: r for r in account}           # ACCOID -> account
mast_by_vch  = {r["PADVCHNO"]: r for r in paidmast}        # PADVCHNO -> paidmast
tran_by_purvch = defaultdict(list)                          # PURVCHNO -> [paidtran rows]
tran_by_padvch = defaultdict(list)                          # PADVCHNO -> [paidtran rows]

for r in paidtran:
    tran_by_purvch[r["PURVCHNO"]].append(r)
    tran_by_padvch[r["PADVCHNO"]].append(r)

# CHQ lookup: integer cheque number -> paidmast row
chq_lookup = {}
for r in paidmast:
    chq = r["CHQNO"].strip()
    if chq.isdigit():
        chq_lookup[int(chq)] = r


# ─── Export Joined DBF to Excel ───────────────────────────────────────────────

print("\nExporting joined DBF data to DBF_Joined_Data.xlsx ...")

wb_dbf = openpyxl.Workbook()

# Sheet 1: PAIDMAST + ACCOUNT join
ws1 = wb_dbf.active
ws1.title = "PAIDMAST_with_Party"

hdr1 = [
    "VoucherNo", "VoucherDate", "PartyID", "PartyName", "Amount",
    "Discount", "Pending", "CHQNo", "CHQDate", "PayType", "Narration", "FYear",
]
ws1.append(hdr1)

for r in paidmast:
    acct = acct_by_id.get(r["ACCOID"], {})
    party_name = acct.get("ACCONM", "")
    vch_date = parse_date(r["PADVCHDATE"])
    chq_date = parse_date(r["CHQDATE"])
    ws1.append([
        int(r["PADVCHNO"]) if r["PADVCHNO"].isdigit() else r["PADVCHNO"],
        vch_date,
        int(r["ACCOID"]) if r["ACCOID"].isdigit() else r["ACCOID"],
        party_name,
        float(r["PADVCHAMT"]) if r["PADVCHAMT"] else 0,
        float(r["PADDISC"]) if r["PADDISC"] else 0,
        float(r["PENDING"]) if r["PENDING"] else 0,
        r["CHQNO"],
        chq_date,
        r["PADTRDTYPE"],
        r["NARA"],
        r["FINYEAR"],
    ])

# Sheet 2: PAIDTRAN + PAIDMAST + ACCOUNT join
ws2 = wb_dbf.create_sheet("PAIDTRAN_Full")
hdr2 = [
    "PayVoucherNo", "PayDate", "PartyID", "PartyName",
    "BillNo", "BillDate", "BillAmt", "PaidAmt", "BalAmt",
    "Discount", "PurchVoucherNo", "CHQNo", "CHQDate", "PayType", "PriorToTakeover",
]
ws2.append(hdr2)

for r in paidtran:
    mast = mast_by_vch.get(r["PADVCHNO"], {})
    acct = acct_by_id.get(r["ACCOID"], {})
    party_name = acct.get("ACCONM", "")
    bill_date = parse_date(r["BILLDATE"])
    pay_date  = parse_date(r["PADVCHDATE"])
    chq_date  = parse_date(mast.get("CHQDATE", ""))
    prior = "YES" if (bill_date and bill_date < TAKEOVER_DT) else "NO"
    ws2.append([
        int(r["PADVCHNO"]) if r["PADVCHNO"].isdigit() else r["PADVCHNO"],
        pay_date,
        int(r["ACCOID"]) if r["ACCOID"].isdigit() else r["ACCOID"],
        party_name,
        r["BILLNO"],
        bill_date,
        float(r["PURNETAMT"]) if r["PURNETAMT"] else 0,
        float(r["PURPAIDAMT"]) if r["PURPAIDAMT"] else 0,
        float(r["PURBALAMT"]) if r["PURBALAMT"] else 0,
        float(r["DISCOUNT"]) if r["DISCOUNT"] else 0,
        int(r["PURVCHNO"]) if r["PURVCHNO"].isdigit() else r["PURVCHNO"],
        mast.get("CHQNO", ""),
        chq_date,
        r["PADTRDTYPE"],
        prior,
    ])

# Style headers
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for ws in [ws1, ws2]:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = ws["A2"]

wb_dbf.save("DBF_Joined_Data.xlsx")
print("  Saved: DBF_Joined_Data.xlsx")


# ─── Step 2 & 3: Update Old_Bank_Statement ───────────────────────────────────

print("\nStep 2 & 3: Updating Old_Bank_Statement_Shree_Seva_Medical.xlsx ...")

wb_bank = openpyxl.load_workbook("Old_Bank_Statement_Shree_Seva_Medical.xlsx")
ws_bank = wb_bank.active

# Find or add Voucher No. and Party Name columns
bank_headers = [cell.value for cell in ws_bank[1]]
print(f"  Existing columns: {bank_headers}")

# Determine column positions
if "Voucher_No" not in bank_headers:
    vch_col = ws_bank.max_column + 1
    ws_bank.cell(row=1, column=vch_col, value="Voucher_No")
else:
    vch_col = bank_headers.index("Voucher_No") + 1

if "Party_Name" not in bank_headers:
    pty_col = ws_bank.max_column + 1
    ws_bank.cell(row=1, column=pty_col, value="Party_Name")
else:
    pty_col = bank_headers.index("Party_Name") + 1

if "Bill_Nos" not in bank_headers:
    bill_col = ws_bank.max_column + 1
    ws_bank.cell(row=1, column=bill_col, value="Bill_Nos")
else:
    bill_col = bank_headers.index("Bill_Nos") + 1

# Style new header cells
new_hdr_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
new_hdr_font = Font(color="FFFFFF", bold=True)
for col in [vch_col, pty_col, bill_col]:
    c = ws_bank.cell(row=1, column=col)
    c.fill = new_hdr_fill
    c.font = new_hdr_font

# Narration column index (col B = index 2)
nar_col_idx = 2  # 1-based

matched_count = 0
for row_idx in range(2, ws_bank.max_row + 1):
    narration = ws_bank.cell(row=row_idx, column=nar_col_idx).value
    if not narration:
        continue
    narration = str(narration)

    # Extract last number from narration (cheque number)
    nums = re.findall(r"\d+", narration)
    if not nums:
        continue
    last_num = int(nums[-1])

    mast_row = chq_lookup.get(last_num)
    if not mast_row:
        continue

    matched_count += 1

    # Voucher No. = payment voucher number
    vch_no = mast_row["PADVCHNO"]
    ws_bank.cell(row=row_idx, column=vch_col, value=int(vch_no) if vch_no.isdigit() else vch_no)

    # Party Name from ACCOUNT
    acct = acct_by_id.get(mast_row["ACCOID"], {})
    party_name = acct.get("ACCONM", "")
    ws_bank.cell(row=row_idx, column=pty_col, value=party_name)

    # Bill Nos - extract from NARA field (split by "." to get bill numbers after "BILLNO.")
    nara = mast_row.get("NARA", "")
    bill_nos = ""
    if "BILLNO" in nara.upper():
        parts = re.split(r"BILLNO\.?", nara, flags=re.IGNORECASE)
        if len(parts) > 1:
            bill_nos = parts[1].strip()
    elif "BILL NO" in nara.upper():
        parts = re.split(r"BILL NO\.?", nara, flags=re.IGNORECASE)
        if len(parts) > 1:
            bill_nos = parts[1].strip()

    # Also get bill numbers from PAIDTRAN
    tran_rows = tran_by_padvch.get(mast_row["PADVCHNO"], [])
    if tran_rows:
        tran_bills = "+".join(r["BILLNO"] for r in tran_rows if r["BILLNO"])
        if bill_nos:
            bill_nos = bill_nos + " | " + tran_bills
        else:
            bill_nos = tran_bills

    ws_bank.cell(row=row_idx, column=bill_col, value=bill_nos[:200] if bill_nos else "")

    # Highlight matched rows
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws_bank.cell(row=row_idx, column=vch_col).fill = match_fill
    ws_bank.cell(row=row_idx, column=pty_col).fill = match_fill
    ws_bank.cell(row=row_idx, column=bill_col).fill = match_fill

print(f"  Bank statement rows matched with PAIDMAST: {matched_count}")

wb_bank.save("Old_Bank_Statement_Shree_Seva_Medical.xlsx")
print("  Saved: Old_Bank_Statement_Shree_Seva_Medical.xlsx")


# ─── Step 4 & 5: Update System_Purchase ──────────────────────────────────────

print("\nStep 4 & 5: Updating System_Puchase.xlsx ...")

wb_sys = openpyxl.load_workbook("System_Puchase.xlsx")
ws_sys = wb_sys.active

sys_headers = [cell.value for cell in ws_sys[1]]
print(f"  Existing columns: {sys_headers}")

# Column indices (1-based)
def col_idx(header, headers):
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == header.lower():
            return i + 1
    return None

voutype_col  = col_idx("Vou.Type ", sys_headers) or col_idx("Vou.Type", sys_headers)
voudate_col  = col_idx("Vou.Date", sys_headers)
vouno_col    = 1  # Vou.No. is column A
date_col     = col_idx("Date", sys_headers)
type_col     = col_idx("Type", sys_headers)
chqno_col    = col_idx("ChQ No.", sys_headers)
prev_pend_col = col_idx("Previous Pending", sys_headers)
old_acct_col  = col_idx("Old Account", sys_headers)
new_acct_col  = col_idx("New Account", sys_headers)
pending_col   = col_idx("Pending", sys_headers)
amount_col    = col_idx("Amount", sys_headers)

print(f"  Column map: Date={date_col}, Type={type_col}, CHQ={chqno_col}, PrevPend={prev_pend_col}")

# Pre-calculate: for each PADVCHNO, sum of PURPAIDAMT before/after takeover
prev_pend_by_padvch = defaultdict(float)  # padvchno -> sum before takeover
new_amt_by_padvch   = defaultdict(float)  # padvchno -> sum on/after takeover

for r in paidtran:
    bill_date = parse_date(r["BILLDATE"])
    paid_amt = float(r["PURPAIDAMT"]) if r["PURPAIDAMT"] else 0
    if bill_date and bill_date < TAKEOVER_DT:
        prev_pend_by_padvch[r["PADVCHNO"]] += paid_amt
    else:
        new_amt_by_padvch[r["PADVCHNO"]] += paid_amt

# Style for updated cells
update_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
match_fill2 = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

sys_matched = 0
sys_prev_pend = 0

for row_idx in range(2, ws_sys.max_row + 1):
    vou_no_cell = ws_sys.cell(row=row_idx, column=vouno_col).value
    if vou_no_cell is None:
        continue
    vou_no_str = str(int(vou_no_cell)) if isinstance(vou_no_cell, (int, float)) else str(vou_no_cell).strip()

    # Find PAIDTRAN rows where PURVCHNO = vou_no
    tran_rows = tran_by_purvch.get(vou_no_str, [])
    if not tran_rows:
        continue

    sys_matched += 1

    # Get payment info from PAIDMAST (via PADVCHNO)
    # A purchase bill may be paid by one payment voucher
    padvchno = tran_rows[0]["PADVCHNO"]  # use first match
    mast_row = mast_by_vch.get(padvchno, {})

    if mast_row:
        chq_no   = mast_row.get("CHQNO", "")
        chq_date = parse_date(mast_row.get("CHQDATE", ""))
        pay_type = mast_row.get("PADTRDTYPE", "")

        # Map type: BKP = Bank Payment
        if chq_no.upper() in ("GPAY", "NEFT", "RTGS", "UPI", "CASH"):
            display_type = chq_no.upper()
            chq_no_display = ""
        else:
            display_type = "CHQ"
            chq_no_display = chq_no

        # Fill Date column
        if date_col:
            cell = ws_sys.cell(row=row_idx, column=date_col, value=chq_date)
            cell.fill = update_fill

        # Fill Type column
        if type_col:
            cell = ws_sys.cell(row=row_idx, column=type_col, value=display_type)
            cell.fill = update_fill

        # Fill CHQ No. column
        if chqno_col:
            cell = ws_sys.cell(row=row_idx, column=chqno_col, value=chq_no_display)
            cell.fill = update_fill

    # Previous Pending = sum of payments for bills BEFORE takeover date (same payment voucher)
    prev_amt = prev_pend_by_padvch.get(padvchno, 0)
    if prev_amt > 0 and prev_pend_col:
        cell = ws_sys.cell(row=row_idx, column=prev_pend_col, value=round(prev_amt, 2))
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        sys_prev_pend += 1

print(f"  System_Purchase rows matched: {sys_matched}")
print(f"  Rows with Previous Pending amount: {sys_prev_pend}")

wb_sys.save("System_Puchase.xlsx")
print("  Saved: System_Puchase.xlsx")


# ─── Summary Report ──────────────────────────────────────────────────────────

print("\n" + "=" * 60)
print("RECONCILIATION SUMMARY")
print("=" * 60)
print(f"  Takeover Date          : 29-May-2025")
print(f"  ACCOUNT records        : {len(account)}")
print(f"  PAIDMAST records       : {len(paidmast)}")
print(f"  PAIDTRAN records       : {len(paidtran)}")
print(f"  Bank Statement rows    : {ws_bank.max_row - 1}")
print(f"  Bank rows matched (CHQ): {matched_count}")
print(f"  System_Purchase rows   : {ws_sys.max_row - 1}")
print(f"  System rows matched    : {sys_matched}")
print(f"  Rows with Prev Pending : {sys_prev_pend}")
print()

total_prev = sum(prev_pend_by_padvch.values())
total_new  = sum(new_amt_by_padvch.values())
print(f"  Total paid (BEFORE 29-May-2025) : Rs. {total_prev:,.2f}  [Previous Pending]")
print(f"  Total paid (ON/AFTER 29-May-2025): Rs. {total_new:,.2f}  [New Bills]")
print()
print("Output files:")
print("  1. DBF_Joined_Data.xlsx              - 3 tables joined")
print("  2. Old_Bank_Statement_Shree_Seva_Medical.xlsx - updated with Voucher_No, Party_Name, Bill_Nos")
print("  3. System_Puchase.xlsx               - updated with CHQ No., Date, Type, Previous Pending")
print("=" * 60)
